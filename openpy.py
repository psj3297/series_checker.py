import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# 새 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "학습 로그"

# 헤더 설정 (A1~I1)
headers = [
    '날짜', '코드 트레이닝(T/F)', 'EXP', '알고리즘 이해(T/F)', 'EXP',
    '복습 노트(T/F)', 'EXP', 'EXP 합계', '메모'
]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)

# 데이터 입력 (대문자 'TRUE'/'FALSE' 문자열로 저장, EXP는 수식)
data = [
    # 행 2: 11/6, all TRUE
    ['11/6', 'TRUE', '=IF(B2="TRUE",15,0)', 'TRUE', '=IF(D2="TRUE",20,0)', 'TRUE', '=IF(F2="TRUE",15,0)', '=C2+E2+G2', 'BFS 학습 완료'],
    # 행 3: 11/7, code FALSE, algo TRUE, review FALSE
    ['11/7', 'FALSE', '=IF(B3="TRUE",15,0)', 'TRUE', '=IF(D3="TRUE",20,0)', 'FALSE', '=IF(F3="TRUE",15,0)', '=C3+E3+G3', '집중 2라운드 성공'],
    # 행 4: 11/8, code TRUE, algo FALSE, review TRUE
    ['11/8', 'TRUE', '=IF(B4="TRUE",15,0)', 'FALSE', '=IF(D4="TRUE",20,0)', 'TRUE', '=IF(F4="TRUE",15,0)', '=C4+E4+G4', '']
]

for row_idx, row_data in enumerate(data, start=2):
    for col, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col, value=value)

# 체크박스 열(B, D, F)에 드롭다운 추가 (TRUE/FALSE 대문자)
dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
dv.errorTitle = '잘못된 값'
dv.error = 'TRUE 또는 FALSE만 선택하세요.'
dv.promptTitle = '체크 선택'
dv.prompt = 'TRUE(체크) 또는 FALSE(비체크)를 선택하세요.'
ws.add_data_validation(dv)
dv.add('B2:B100')  # 100행까지 (충분함)
dv.add('D2:D100')
dv.add('F2:F100')

# 시각적 효과: TRUE일 때 셀 배경을 녹색으로 (조건부 서식)
green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
dxf = DifferentialStyle(fill=green_fill)

# B 열 규칙
true_rule_b = Rule(type="expression", formula=['$B2="TRUE"'], dxf=dxf)
ws.conditional_formatting.add('B2:B100', true_rule_b)

# D 열 규칙
true_rule_d = Rule(type="expression", formula=['$D2="TRUE"'], dxf=dxf)
ws.conditional_formatting.add('D2:D100', true_rule_d)

# F 열 규칙
true_rule_f = Rule(type="expression", formula=['$F2="TRUE"'], dxf=dxf)
ws.conditional_formatting.add('F2:F100', true_rule_f)

# 열 너비 조정
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 8
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 20

# 파일 저장
wb.save('training_log_final.xlsx')
print("최종 엑셀 파일 'training_log_final.xlsx'이 생성되었습니다! 열어서 드롭다운 확인하세요.")